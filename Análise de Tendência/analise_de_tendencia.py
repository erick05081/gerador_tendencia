# importando bibliotecas necessárias
from typing import Optional, Iterable, List
import re
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

# ==========================
# Configurações iniciais
# ==========================
st.set_page_config(page_title="Análise de Tendência")

pd.set_option(
    "display.float_format",
    lambda v: f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
)

st.title("Gerador de Análise de Tendência")
st.write("Este aplicativo gera análises de tendência com base em dados fornecidos pelo usuário.")
st.subheader("Enviar arquivos:")

# ==========================
# Carregamento do cadastro de projetos (df_cc)
# ==========================
df_cc = None
CAMINHO_DF_CC = (
    "https://raw.githubusercontent.com/"
    "erick05081/gerador_tendencia/main/"
    "data/pacaembu_base_cadastro_projetos_ajustado_final.xlsx"
)
try:
    df_cc = pd.read_excel(CAMINHO_DF_CC, sheet_name="Projeto")
except Exception:
    st.warning(
        "Não foi possível carregar o cadastro de projetos (df_cc) pelo caminho padrão. "
        "Você ainda pode rodar sem 'des. imobiliario' (automático) ou subir o cadastro abaixo."
    )
    df_cc_file = st.file_uploader(
        "Opcional: envie o cadastro de projetos (df_cc) com as colunas ['codigo','des. imobiliario']",
        type=["xlsx"],
        key="upload_df_cc"
    )
    if df_cc_file:
        try:
            df_cc = pd.read_excel(df_cc_file, sheet_name="Projeto")
            st.success("Cadastro df_cc carregado com sucesso via upload.")
        except Exception as e2:
            st.error(f"Falha ao ler o df_cc enviado: {e2}")
            df_cc = None

# ==========================
# Funções utilitárias
# ==========================
def parse_first_date(df: pd.DataFrame, date_col: str = "Data Base"):
    """Tenta ler a data da primeira linha da coluna informada."""
    if date_col not in df.columns:
        raise KeyError(f"Coluna '{date_col}' não encontrada.")
    first_val = df[date_col].dropna().head(1)
    if first_val.empty:
        return pd.NaT
    return pd.to_datetime(first_val.iloc[0], errors="coerce", dayfirst=True)


def carregar_dfs_ordenados_por_data(uploaded_files, sheet_name="BD Acomp", date_col="Data Base"):
    """
    Lê os arquivos, coleta a data da primeira linha da coluna `date_col`,
    e retorna (df_mais_recente, df_mais_antigo, detalhes).
    """
    registros = []

    for uploaded_file in uploaded_files:
        try:
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
        except ValueError as e:
            st.warning(f"Arquivo '{uploaded_file.name}': não encontrei a planilha '{sheet_name}'. Detalhe: {e}")
            continue
        except Exception as e:
            st.warning(f"Falha ao ler '{uploaded_file.name}': {e}")
            continue

        try:
            first_date = parse_first_date(df, date_col=date_col)
        except KeyError as e:
            st.warning(f"Arquivo '{uploaded_file.name}': {e}")
            first_date = pd.NaT

        registros.append({
            "file": uploaded_file,
            "name": getattr(uploaded_file, "name", "arquivo_sem_nome.xlsx"),
            "df": df,
            "date": first_date
        })

    if not registros:
        return None, None, []

    validos   = [r for r in registros if pd.notna(r["date"])]
    invalidos = [r for r in registros if pd.isna(r["date"])]

    if not validos:
        st.info("Nenhum arquivo possui uma data válida na primeira linha de 'Data Base'.")
        return None, None, registros

    validos_sorted = sorted(validos, key=lambda r: r["date"])
    mais_antigo  = validos_sorted[0]
    mais_recente = validos_sorted[-1]

    df1 = mais_recente["df"]  # mais recente
    df2 = mais_antigo["df"]   # mais antigo

    return df1, df2, validos_sorted + invalidos


# Dicionário de mapeamento (colunas originais -> novas colunas)
COLS_ORIGINAIS = ["C.C.", "Nome", "U.C.", "Realizado", "A Incorrer", "A Contratar", "Distrato"]
MAPA_RENOME = {
    "C.C.": "codigo",
    "Nome": "nome",
    "U.C.": "uc",
    "Realizado": "realizado",
    "A Incorrer": "incorrer",
    "A Contratar": "contratar",
    "Distrato": "distrato",
}

def padronizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    faltantes = [c for c in COLS_ORIGINAIS if c not in df.columns]
    if faltantes:
        raise KeyError(f"Colunas obrigatórias não encontradas: {faltantes}")
    return df[COLS_ORIGINAIS].rename(columns=MAPA_RENOME)

def transformar_df1_df2(df1: pd.DataFrame, df2: pd.DataFrame):
    df1_proc = padronizar_colunas(df1)
    df2_proc = padronizar_colunas(df2)
    return df1_proc, df2_proc

def padroniza_codigo(df, coluna="codigo"):
    """Mantém apenas os 3 últimos dígitos quando o valor tiver mais de 3 caracteres."""
    df = df.copy()
    df[coluna] = df[coluna].astype(str).apply(lambda x: x if len(x) <= 3 else x[-3:])
    return df

# -------- Ordenação natural das UCs --------
def natural_key(s: str):
    """Quebra a string em blocos de dígitos e não-dígitos para ordenar naturalmente."""
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r"(\d+)", str(s))]

def ordenar_uc_natural(columns: Iterable) -> List:
    """
    Ordena colunas de UC de forma natural:
    - Se todos forem numéricos, ordena numericamente.
    - Se houver misto (texto e números), usa ordenação natural (1,2,3,...,9,10,11, Apto, Casa, ...).
    """
    cols = list(columns)
    numericos, nao_numericos = [], []
    for c in cols:
        try:
            numericos.append((c, int(str(c))))
        except Exception:
            nao_numericos.append(c)

    if len(numericos) == len(cols):
        return [c for c, _ in sorted(numericos, key=lambda x: x[1])]
    return sorted(cols, key=natural_key)

# -------- Construção de matriz por UC --------
def matriz_por_uc(
    df: pd.DataFrame,
    use_des_imobiliario: bool = True,
    df_cc: Optional[pd.DataFrame] = None,
    cols_status: Optional[list[str]] = None,
    round_ndigits: int = 2,
) -> pd.DataFrame:
    """
    Constrói uma matriz com índice = 'nome' (e 'des. imobiliario' se use_des_imobiliario=True),
    colunas = valores da coluna 'uc',
    valores = soma de (realizado + incorrer + contratar + distrato) por (nome[, des. imobiliario], uc).
    """
    if cols_status is None:
        cols_status = ["realizado", "incorrer", "contratar", "distrato"]

    for col_obr in ["nome", "uc"]:
        if col_obr not in df.columns:
            raise ValueError(f"Coluna obrigatória ausente: '{col_obr}'")

    work = df.copy()

    # Normaliza UC para evitar duplicidades por caixa/espaços
    work["uc"] = work["uc"].astype(str).str.strip()

    # Garante 'des. imobiliario' se for solicitado
    if use_des_imobiliario and "des. imobiliario" not in work.columns:
        if df_cc is None or "codigo" not in work.columns or "codigo" not in df_cc.columns:
            raise ValueError("Para usar 'des. imobiliario', informe df_cc e garanta 'codigo' em ambos.")
        tmp_cc = df_cc[["codigo", "des. imobiliario"]].copy()
        tmp_cc["codigo"] = tmp_cc["codigo"].astype(str)
        work["codigo"] = work["codigo"].astype(str)
        work = work.merge(tmp_cc, on="codigo", how="left")
        work["des. imobiliario"] = work["des. imobiliario"].fillna("")

    # Converte SOMENTE as colunas de status para numérico
    present = [c for c in cols_status if c in work.columns]
    if not present:
        raise ValueError("Nenhuma coluna de status encontrada no DataFrame.")
    work[present] = work[present].apply(pd.to_numeric, errors="coerce").fillna(0)

    # Total por linha
    work["Total_Item"] = work[present].sum(axis=1, skipna=True)

    # Agrupa e pivota
    if use_des_imobiliario:
        agg = work.groupby(["nome", "des. imobiliario", "uc"], as_index=False)["Total_Item"].sum()
        matriz = agg.pivot(index=["nome", "des. imobiliario"], columns="uc", values="Total_Item")
    else:
        agg = work.groupby(["nome", "uc"], as_index=False)["Total_Item"].sum()
        matriz = agg.pivot(index="nome", columns="uc", values="Total_Item")

    matriz = matriz.fillna(0.0).astype(float).round(round_ndigits)

    # Ordena as colunas de UC de forma natural/numérica
    ordem_uc = ordenar_uc_natural(matriz.columns)
    matriz = matriz.reindex(columns=ordem_uc)

    # Mantém 'des. imobiliario' como PRIMEIRA coluna após 'nome'
    matriz = matriz.reset_index()
    if "des. imobiliario" in matriz.columns and "nome" in matriz.columns:
        col = matriz.pop("des. imobiliario")
        idx = matriz.columns.get_loc("nome") + 1
        matriz.insert(idx, "des. imobiliario", col)
    matriz = matriz.set_index("nome").sort_index()

    return matriz

# -------- Delta por UC --------
def delta_por_uc(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    use_des_imobiliario: bool = True,
    df_cc: Optional[pd.DataFrame] = None,
    add_total: bool = False,
    round_ndigits: int = 2,
) -> pd.DataFrame:
    """
    Calcula o delta por UC: (matriz_por_uc(df1) - matriz_por_uc(df2)).
    Alinha índices e colunas (união) e preenche ausências com 0.
    """
    m1 = matriz_por_uc(df1, use_des_imobiliario=use_des_imobiliario, df_cc=df_cc, round_ndigits=round_ndigits)
    m2 = matriz_por_uc(df2, use_des_imobiliario=use_des_imobiliario, df_cc=df_cc, round_ndigits=round_ndigits)

    # União de linhas e colunas
    idx_union = m1.index.union(m2.index)
    col_union = m1.columns.union(m2.columns)

    # Colunas não numéricas (ex.: 'des. imobiliario')
    non_numeric_cols = [c for c in col_union
                        if (c in m1.columns and m1[c].dtype == "O") or (c in m2.columns and m2[c].dtype == "O")]

    # Partes numéricas reindexadas (ordenadas naturalmente)
    m1_num = m1.drop(columns=[c for c in m1.columns if c in non_numeric_cols], errors="ignore")
    m2_num = m2.drop(columns=[c for c in m2.columns if c in non_numeric_cols], errors="ignore")
    num_cols_union = [c for c in col_union if c not in non_numeric_cols]
    num_cols_union = ordenar_uc_natural(num_cols_union)

    m1_num = m1_num.reindex(index=idx_union, columns=num_cols_union, fill_value=0.0)
    m2_num = m2_num.reindex(index=idx_union, columns=num_cols_union, fill_value=0.0)

    # Delta numérico
    out = (m1_num - m2_num).round(round_ndigits)

    # Reinsere 'des. imobiliario' como PRIMEIRA coluna após 'nome'
    has_des = ("des. imobiliario" in non_numeric_cols)
    if has_des:
        out = out.reset_index()
        base_non_num = m1.reindex(idx_union)[["des. imobiliario"]] if "des. imobiliario" in m1.columns \
                       else m2.reindex(idx_union)[["des. imobiliario"]]
        out.insert(out.columns.get_loc("nome") + 1, "des. imobiliario", base_non_num["des. imobiliario"].values)
        out = out.set_index("nome")

    # (opcional) Delta_Total (soma horizontal)
    #if add_total and not out.empty:
        #num_cols_final = out.select_dtypes(include=[float, int]).columns
        #out["Delta_Total"] = out[num_cols_final].sum(axis=1).round(round_ndigits)

    # >>> ORDEM FINAL DAS COLUNAS <<<
    # des. imobiliario (se existir) -> UCs (ordem natural) -> Delta_Total (se existir)
    numeric_cols_final = [c for c in out.columns if c not in ["des. imobiliario", "Delta_Total"]]
    numeric_cols_final = ordenar_uc_natural(numeric_cols_final)

    cols_final = []
    if has_des:
        cols_final.append("des. imobiliario")
    cols_final += numeric_cols_final
    if "Delta_Total" in out.columns:
        cols_final.append("Delta_Total")

    out = out.reindex(columns=cols_final).sort_index()
    return out

# ==========================
# Estado (session_state)
# ==========================
if "df1" not in st.session_state:
    st.session_state.df1 = None
if "df2" not in st.session_state:
    st.session_state.df2 = None

# ==========================
# UI Streamlit (upload)
# ==========================
uploaded_files = st.file_uploader(
    "Requer exatamente 2 arquivos .xlsx",
    accept_multiple_files=True,
    type="xlsx"
)

if uploaded_files:
    if len(uploaded_files) != 2:
        st.error("Por favor, envie exatamente 2 arquivos.")
        st.stop()

    df1_tmp, df2_tmp, detalhe = carregar_dfs_ordenados_por_data(
        uploaded_files,
        sheet_name="BD Acomp",
        date_col="Data Base"
    )

    # Guarda no estado
    st.session_state.df1 = df1_tmp
    st.session_state.df2 = df2_tmp

# Recupera do estado
df1 = st.session_state.df1
df2 = st.session_state.df2

# ==========================
# Processamento com barra de progresso + Download
# ==========================
if df1 is not None and df2 is not None:
    # Barra de progresso e status
    progress = st.progress(0)
    status = st.empty()

    try:
        # 1) Padroniza as colunas
        status.markdown("**Padronizando colunas...**")
        progress.progress(10)
        df1, df2 = transformar_df1_df2(df1, df2)

        # 2) Padroniza códigos
        status.markdown("**Padronizando códigos...**")
        progress.progress(20)
        df1 = padroniza_codigo(df1, "codigo")
        df2 = padroniza_codigo(df2, "codigo")

        # 3) Delta por UC
        status.markdown("**Calculando Delta por UC...**")
        progress.progress(45)
        use_des = df_cc is not None
        delta_uc_df = delta_por_uc(
            df1,
            df2,
            use_des_imobiliario=use_des,
            df_cc=df_cc,
            add_total=True
        )

        # 4) Estilização
        status.markdown("**Aplicando estilos...**")
        progress.progress(65)

        def zebra_mask(df: pd.DataFrame, color="#e6f5ff", one_based=True):
            nrows, ncols = df.shape
            row_pos = np.arange(nrows)
            rows_to_paint = (row_pos % 2 == 0) if one_based else (row_pos % 2 == 1)
            mask = np.full((nrows, ncols), "", dtype=object)
            mask[rows_to_paint, :] = f"background-color: {color}"
            return pd.DataFrame(mask, index=df.index, columns=df.columns)

        delta_uc_styled = (
            delta_uc_df.style
                .apply(zebra_mask, axis=None, color="#e6f5ff", one_based=True)
                .set_table_styles(
                    [
                        {
                            "selector": "th",
                            "props": [
                                ("background-color", "#E6F0FF"),
                                ("font-weight", "bold"),
                                ("text-align", "center"),
                            ],
                        },
                    ]
                )
                .set_properties(
                    **{
                        "border-color": "#D9D9D9",
                        "border-style": "solid",
                        "border-width": "0.5pt",
                        "text-align": "center",
                    }
                )
        )

        # 5) Geração do Excel em memória
        status.markdown("**Gerando arquivo Excel em memória...**")
        progress.progress(80)
        output_buffer = BytesIO()
        sheet = "Delta_por_UC"

        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            delta_uc_styled.to_excel(writer, sheet_name=sheet)

        # 6) Ajustes finais com openpyxl em memória
        status.markdown("**Aplicando ajustes finais no Excel...**")
        progress.progress(90)
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        output_buffer.seek(0)
        wb = load_workbook(output_buffer)
        ws = wb[sheet]
        max_row = ws.max_row
        max_col = ws.max_column

        # Centraliza cabeçalho
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Centraliza corpo e define formato numérico onde aplicável
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        # Autoajuste de colunas
        PADDING = 2.0
        MAX_WIDTH = 60.0
        MIN_WIDTH = 8.5

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value is None:
                    cell_len = 0
                else:
                    text = str(cell.value).replace("\r", "")
                    cell_len = max(len(line) for line in text.split("\n"))
                if isinstance(cell.value, (int, float)):
                    cell_len += 1
                max_len = max(max_len, cell_len)

            adjusted = max(max_len + PADDING, MIN_WIDTH)
            adjusted = min(adjusted, MAX_WIDTH)
            ws.column_dimensions[col_letter].width = adjusted

        # Salva o workbook de volta em memória
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        # 7) Finaliza progresso e mostra botão de download
        progress.progress(100)
        status.markdown("✅ **Pronto! Clique para baixar o arquivo.**")

        st.download_button(
            "Baixar Excel (Delta por UC)",
            data=final_buffer.getvalue(),
            file_name="Análise de Tendência por Projeto - Delta.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except KeyError as e:
        status.markdown("❌ Ocorreu um erro durante o processamento.")
        st.error(f"Erro ao padronizar/calcular delta: {e}")
        progress.progress(0)
    except Exception as e:
        status.markdown("❌ Ocorreu um erro durante o processamento.")
        st.error(f"Ocorreu um erro: {e}")
        progress.progress(0)
else:
    st.info("Envie 2 arquivos .xlsx para iniciar a análise.")
